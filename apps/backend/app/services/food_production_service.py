import io
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import date
from decimal import Decimal

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.exc import IntegrityError
from sqlmodel import Session, select

from app.models.admission import Admission
from app.models.allergy import PatientAllergyIntolerance
from app.models.care_unit import CareUnit
from app.models.common import utc_now
from app.models.food_production import (
    MEAL_TIMES,
    FoodRegimenCatalogItem,
    NutritionalMealPlan,
    NutritionalMealPlanItem,
    NutritionalMealPlanSlot,
    NutritionalModularPreparation,
)
from app.models.hospital_service import HospitalService
from app.models.patient import Patient
from app.models.patient_location_history import PatientLocationHistory
from app.models.room import Room
from app.schemas.food_production import (
    FoodCatalogItemCreate,
    FoodCatalogItemRead,
    FoodCatalogItemUpdate,
    MealPlanCreate,
    MealPlanItemRead,
    MealPlanRead,
    MealPlanSlotRead,
    MealPlanUpdate,
    ModularPreparationRead,
    ProductionConsolidatedRead,
    ProductionException,
    ProductionModularDetail,
    ProductionPreparationLine,
    ProductionRationDetail,
    ProductionSummaryLine,
)
from app.services.audit_service import record_audit


MEAL_LABELS = {
    "breakfast": "Desayuno",
    "morning_snack": "Colación AM",
    "lunch": "Almuerzo",
    "afternoon_snack": "Once",
    "dinner": "Cena",
    "night_snack": "Colación PM",
}


def _not_found(label: str) -> HTTPException:
    return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"{label} no encontrado.")


def _conflict(detail: str) -> HTTPException:
    return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=detail)


def _commit(session: Session) -> None:
    try:
        session.commit()
    except IntegrityError as exc:
        session.rollback()
        raise _conflict("No fue posible guardar la minuta por un conflicto de datos.") from exc


def normalize_catalog_name(value: str) -> str:
    text = unicodedata.normalize("NFD", value.casefold())
    text = "".join(character for character in text if unicodedata.category(character) != "Mn")
    return " ".join(text.split())


def catalog_code(value: str) -> str:
    normalized = normalize_catalog_name(value)
    code = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return code[:70] or "item"


def list_catalog(
    session: Session, *, search: str | None = None, include_inactive: bool = False
) -> list[FoodCatalogItemRead]:
    statement = select(FoodRegimenCatalogItem)
    if not include_inactive:
        statement = statement.where(FoodRegimenCatalogItem.is_active.is_(True))
    if search:
        statement = statement.where(
            FoodRegimenCatalogItem.normalized_name.contains(normalize_catalog_name(search))
        )
    rows = session.exec(statement.order_by(FoodRegimenCatalogItem.display_name)).all()
    return [FoodCatalogItemRead.model_validate(row) for row in rows]


def create_catalog_item(
    session: Session, payload: FoodCatalogItemCreate, actor_id: uuid.UUID
) -> FoodCatalogItemRead:
    normalized = normalize_catalog_name(payload.display_name)
    existing = session.exec(
        select(FoodRegimenCatalogItem).where(
            FoodRegimenCatalogItem.normalized_name == normalized
        )
    ).first()
    if existing:
        raise _conflict("Ya existe un ítem equivalente en el catálogo.")
    base_code = catalog_code(payload.display_name)
    code = base_code
    suffix = 2
    while session.exec(
        select(FoodRegimenCatalogItem.id).where(FoodRegimenCatalogItem.code == code)
    ).first():
        code = f"{base_code[:65]}-{suffix}"
        suffix += 1
    row = FoodRegimenCatalogItem(
        code=code,
        normalized_name=normalized,
        **payload.model_dump(),
    )
    session.add(row)
    session.flush()
    record_audit(
        session,
        action="food_catalog_item_created",
        actor_user_id=actor_id,
        entity_type="food_regimen_catalog_item",
        entity_id=row.id,
        after_state={"code": row.code, "item_type": row.item_type, "is_active": True},
    )
    _commit(session)
    session.refresh(row)
    return FoodCatalogItemRead.model_validate(row)


def update_catalog_item(
    session: Session,
    item_id: uuid.UUID,
    payload: FoodCatalogItemUpdate,
    actor_id: uuid.UUID,
) -> FoodCatalogItemRead:
    row = session.get(FoodRegimenCatalogItem, item_id)
    if row is None:
        raise _not_found("Ítem de catálogo")
    normalized = normalize_catalog_name(payload.display_name)
    duplicate = session.exec(
        select(FoodRegimenCatalogItem).where(
            FoodRegimenCatalogItem.normalized_name == normalized,
            FoodRegimenCatalogItem.id != row.id,
        )
    ).first()
    if duplicate:
        raise _conflict("Ya existe un ítem equivalente en el catálogo.")
    before = {"item_type": row.item_type, "is_active": row.is_active}
    for key, value in payload.model_dump().items():
        setattr(row, key, value)
    row.normalized_name = normalized
    row.updated_at = utc_now()
    session.add(row)
    record_audit(
        session,
        action="food_catalog_item_updated",
        actor_user_id=actor_id,
        entity_type="food_regimen_catalog_item",
        entity_id=row.id,
        before_state=before,
        after_state={"item_type": row.item_type, "is_active": row.is_active},
    )
    _commit(session)
    session.refresh(row)
    return FoodCatalogItemRead.model_validate(row)


def _active_admission(session: Session, admission_id: uuid.UUID) -> Admission:
    admission = session.get(Admission, admission_id)
    if admission is None:
        raise _not_found("Hospitalización")
    if admission.status != "active":
        raise _conflict("Un episodio histórico es de solo lectura.")
    return admission


def _replace_plan_children(session: Session, plan: NutritionalMealPlan, payload) -> None:
    slots = session.exec(
        select(NutritionalMealPlanSlot).where(NutritionalMealPlanSlot.meal_plan_id == plan.id)
    ).all()
    for slot in slots:
        for item in session.exec(
            select(NutritionalMealPlanItem).where(
                NutritionalMealPlanItem.meal_plan_slot_id == slot.id
            )
        ).all():
            session.delete(item)
        session.delete(slot)
    for preparation in session.exec(
        select(NutritionalModularPreparation).where(
            NutritionalModularPreparation.meal_plan_id == plan.id
        )
    ).all():
        session.delete(preparation)
    session.flush()

    for slot_input in payload.slots:
        slot_values = slot_input.model_dump(exclude={"items"})
        is_special = bool(
            slot_values["is_special"]
            or slot_values["special_instructions"]
            or any(item.custom_name or item.instructions for item in slot_input.items)
        )
        slot_values["is_special"] = is_special
        slot = NutritionalMealPlanSlot(
            meal_plan_id=plan.id,
            **slot_values,
        )
        session.add(slot)
        session.flush()
        for item_index, item_input in enumerate(slot_input.items):
            if item_input.catalog_item_id is not None:
                catalog = session.get(FoodRegimenCatalogItem, item_input.catalog_item_id)
                if catalog is None or not catalog.is_active:
                    raise _conflict("La minuta contiene un ítem de catálogo inexistente o inactivo.")
            session.add(
                NutritionalMealPlanItem(
                    meal_plan_slot_id=slot.id,
                    sort_order=item_index,
                    **item_input.model_dump(),
                )
            )
    for index, preparation in enumerate(payload.modular_preparations):
        session.add(
            NutritionalModularPreparation(
                meal_plan_id=plan.id,
                sort_order=index,
                **preparation.model_dump(),
            )
        )


def _apply_plan_payload(plan: NutritionalMealPlan, payload, actor_id: uuid.UUID) -> None:
    for field in (
        "effective_from",
        "effective_until",
        "validity_mode",
        "oral_enabled",
        "enteral_enabled",
        "parenteral_enabled",
        "general_instructions",
    ):
        setattr(plan, field, getattr(payload, field))
    plan.updated_by_user_id = actor_id
    plan.updated_at = utc_now()


def create_meal_plan(
    session: Session, admission_id: uuid.UUID, payload: MealPlanCreate, actor_id: uuid.UUID
) -> MealPlanRead:
    _active_admission(session, admission_id)
    existing_draft = session.exec(
        select(NutritionalMealPlan).where(
            NutritionalMealPlan.admission_id == admission_id,
            NutritionalMealPlan.status == "draft",
        )
    ).first()
    if existing_draft:
        raise _conflict("Ya existe una minuta en borrador para esta hospitalización.")
    plan = NutritionalMealPlan(
        admission_id=admission_id,
        effective_from=payload.effective_from,
        created_by_user_id=actor_id,
        updated_by_user_id=actor_id,
    )
    _apply_plan_payload(plan, payload, actor_id)
    session.add(plan)
    session.flush()
    _replace_plan_children(session, plan, payload)
    record_audit(
        session,
        action="meal_plan_draft_created",
        actor_user_id=actor_id,
        entity_type="nutritional_meal_plan",
        entity_id=plan.id,
        admission_id=admission_id,
        after_state={"status": "draft", "version": 1},
    )
    _commit(session)
    return read_meal_plan(session, plan.id)


def update_meal_plan(
    session: Session, plan_id: uuid.UUID, payload: MealPlanUpdate, actor_id: uuid.UUID
) -> MealPlanRead:
    plan = session.get(NutritionalMealPlan, plan_id)
    if plan is None:
        raise _not_found("Minuta")
    _active_admission(session, plan.admission_id)
    if plan.status != "draft":
        raise _conflict("Sólo una minuta en borrador puede modificarse.")
    if plan.version != payload.version:
        raise _conflict("La minuta fue modificada por otro usuario. Recargue los datos.")
    _apply_plan_payload(plan, payload, actor_id)
    plan.version += 1
    session.add(plan)
    _replace_plan_children(session, plan, payload)
    record_audit(
        session,
        action="meal_plan_draft_updated",
        actor_user_id=actor_id,
        entity_type="nutritional_meal_plan",
        entity_id=plan.id,
        admission_id=plan.admission_id,
        before_state={"version": payload.version},
        after_state={"version": plan.version},
    )
    _commit(session)
    return read_meal_plan(session, plan.id)


def finalize_meal_plan(
    session: Session, plan_id: uuid.UUID, version: int, actor_id: uuid.UUID
) -> MealPlanRead:
    plan = session.get(NutritionalMealPlan, plan_id)
    if plan is None:
        raise _not_found("Minuta")
    _active_admission(session, plan.admission_id)
    if plan.status != "draft":
        raise _conflict("La minuta ya no está en borrador.")
    if plan.version != version:
        raise _conflict("La minuta fue modificada por otro usuario. Recargue los datos.")
    plan.status = "finalized"
    plan.version += 1
    plan.finalized_by_user_id = actor_id
    plan.finalized_at = utc_now()
    plan.updated_by_user_id = actor_id
    plan.updated_at = plan.finalized_at
    session.add(plan)
    record_audit(
        session,
        action="meal_plan_finalized",
        actor_user_id=actor_id,
        entity_type="nutritional_meal_plan",
        entity_id=plan.id,
        admission_id=plan.admission_id,
        before_state={"status": "draft", "version": version},
        after_state={"status": "finalized", "version": plan.version},
    )
    _commit(session)
    return read_meal_plan(session, plan.id)


def cancel_meal_plan(
    session: Session, plan_id: uuid.UUID, version: int, actor_id: uuid.UUID
) -> MealPlanRead:
    plan = session.get(NutritionalMealPlan, plan_id)
    if plan is None:
        raise _not_found("Minuta")
    if plan.status != "draft" or plan.version != version:
        raise _conflict("Sólo el borrador vigente puede cancelarse.")
    plan.status = "cancelled"
    plan.version += 1
    plan.updated_by_user_id = actor_id
    plan.updated_at = utc_now()
    session.add(plan)
    record_audit(
        session,
        action="meal_plan_cancelled",
        actor_user_id=actor_id,
        entity_type="nutritional_meal_plan",
        entity_id=plan.id,
        admission_id=plan.admission_id,
        after_state={"status": "cancelled", "version": plan.version},
    )
    _commit(session)
    return read_meal_plan(session, plan.id)


def _item_read(session: Session, item: NutritionalMealPlanItem) -> MealPlanItemRead:
    catalog = session.get(FoodRegimenCatalogItem, item.catalog_item_id) if item.catalog_item_id else None
    return MealPlanItemRead(
        id=item.id,
        catalog_item_id=item.catalog_item_id,
        catalog_code=catalog.code if catalog else None,
        display_name=catalog.display_name if catalog else item.custom_name or "Preparación libre",
        is_custom=catalog is None,
        quantity=item.quantity,
        unit=item.unit,
        instructions=item.instructions,
        sort_order=item.sort_order,
    )


def read_meal_plan(session: Session, plan_id: uuid.UUID) -> MealPlanRead:
    plan = session.get(NutritionalMealPlan, plan_id)
    if plan is None:
        raise _not_found("Minuta")
    slots: list[MealPlanSlotRead] = []
    slot_rows = session.exec(
        select(NutritionalMealPlanSlot)
        .where(NutritionalMealPlanSlot.meal_plan_id == plan.id)
    ).all()
    meal_order = {meal_time: index for index, meal_time in enumerate(MEAL_TIMES)}
    for slot in sorted(slot_rows, key=lambda row: meal_order[row.meal_time]):
        items = session.exec(
            select(NutritionalMealPlanItem)
            .where(NutritionalMealPlanItem.meal_plan_slot_id == slot.id)
            .order_by(NutritionalMealPlanItem.sort_order, NutritionalMealPlanItem.id)
        ).all()
        slots.append(
            MealPlanSlotRead(
                id=slot.id,
                meal_time=slot.meal_time,
                fulfillment_status=slot.fulfillment_status,
                is_special=slot.is_special,
                special_instructions=slot.special_instructions,
                items=[_item_read(session, item) for item in items],
            )
        )
    preparations = session.exec(
        select(NutritionalModularPreparation)
        .where(NutritionalModularPreparation.meal_plan_id == plan.id)
        .order_by(NutritionalModularPreparation.sort_order, NutritionalModularPreparation.id)
    ).all()
    return MealPlanRead(
        **{field: getattr(plan, field) for field in MealPlanRead.model_fields if field not in {"slots", "modular_preparations"}},
        slots=slots,
        modular_preparations=[ModularPreparationRead.model_validate(row) for row in preparations],
    )


def current_meal_plan(
    session: Session, admission_id: uuid.UUID, *, service_date: date | None = None
) -> MealPlanRead | None:
    _active_admission(session, admission_id)
    draft = session.exec(
        select(NutritionalMealPlan)
        .where(
            NutritionalMealPlan.admission_id == admission_id,
            NutritionalMealPlan.status == "draft",
        )
        .order_by(NutritionalMealPlan.updated_at.desc())
    ).first()
    if draft and service_date is None:
        return read_meal_plan(session, draft.id)
    target = service_date or date.today()
    row = session.exec(
        select(NutritionalMealPlan)
        .where(
            NutritionalMealPlan.admission_id == admission_id,
            NutritionalMealPlan.status == "finalized",
            NutritionalMealPlan.effective_from <= target,
            (NutritionalMealPlan.effective_until.is_(None))
            | (NutritionalMealPlan.effective_until >= target),
        )
        .order_by(
            NutritionalMealPlan.effective_from.desc(),
            NutritionalMealPlan.finalized_at.desc(),
        )
    ).first()
    return read_meal_plan(session, row.id) if row else None


def _patient_name(patient: Patient) -> str:
    parts = [patient.given_names, patient.first_surname, patient.second_surname]
    name = " ".join(part for part in parts if part)
    return name or patient.temporary_identifier or patient.hospital_identifier or "Paciente sin nombre"


def _location(session: Session, admission_id: uuid.UUID):
    location = session.exec(
        select(PatientLocationHistory).where(
            PatientLocationHistory.admission_id == admission_id,
            PatientLocationHistory.ended_at.is_(None),
        )
    ).first()
    if not location:
        return None
    unit = session.get(CareUnit, location.care_unit_id)
    room = session.get(Room, unit.room_id) if unit else None
    service = session.get(HospitalService, room.service_id) if room else None
    if not unit or not room or not service:
        return None
    return service, room, unit


def _food_alerts(session: Session, patient_id: uuid.UUID) -> list[str]:
    rows = session.exec(
        select(PatientAllergyIntolerance).where(
            PatientAllergyIntolerance.patient_id == patient_id,
            PatientAllergyIntolerance.category == "food",
            PatientAllergyIntolerance.clinical_status == "active",
            PatientAllergyIntolerance.verification_status.notin_(("refuted", "entered_in_error")),
        )
    ).all()
    return sorted({row.substance_name for row in rows})


def production_consolidated(
    session: Session, *, service_date: date, meal_time: str | None = None
) -> ProductionConsolidatedRead:
    if meal_time is not None and meal_time not in MEAL_TIMES:
        raise HTTPException(status_code=422, detail="Tiempo de comida no válido.")
    summaries: dict[tuple, dict[str, int]] = defaultdict(
        lambda: {"standard": 0, "special": 0, "total": 0}
    )
    item_totals: dict[tuple, Decimal] = defaultdict(Decimal)
    item_patients: dict[tuple, set[uuid.UUID]] = defaultdict(set)
    rations: list[ProductionRationDetail] = []
    modular: list[ProductionModularDetail] = []
    exceptions: list[ProductionException] = []

    admissions = session.exec(
        select(Admission).where(Admission.status == "active").order_by(Admission.admitted_at)
    ).all()
    for admission in admissions:
        patient = session.get(Patient, admission.patient_id)
        if patient is None:
            continue
        patient_name = _patient_name(patient)
        location = _location(session, admission.id)
        if location is None:
            exceptions.append(
                ProductionException(
                    admission_id=admission.id,
                    patient_name=patient_name,
                    service_name=None,
                    room_name=None,
                    bed_name=None,
                    reason="Hospitalización activa sin ubicación vigente.",
                )
            )
            continue
        service, room, unit = location
        plan = current_meal_plan(session, admission.id, service_date=service_date)
        if plan is None:
            exceptions.append(
                ProductionException(
                    admission_id=admission.id,
                    patient_name=patient_name,
                    service_name=service.name,
                    room_name=room.name,
                    bed_name=unit.label or unit.code,
                    reason="Sin minuta finalizada vigente para la fecha.",
                )
            )
            continue
        alerts = _food_alerts(session, patient.id)
        for slot in plan.slots:
            if slot.fulfillment_status != "ordered" or (meal_time and slot.meal_time != meal_time):
                continue
            special = bool(slot.is_special or slot.special_instructions or alerts or any(item.is_custom or item.instructions for item in slot.items))
            key = (service.id, service.name, slot.meal_time)
            item_labels: list[str] = []
            base_ration_count = 0
            for item in slot.items:
                label = f"{item.quantity:g} {item.unit} · {item.display_name}"
                if item.instructions:
                    label += f" ({item.instructions})"
                item_labels.append(label)
                item_key = (service.name, slot.meal_time, item.display_name, item.unit)
                item_totals[item_key] += item.quantity
                item_patients[item_key].add(admission.id)
                if item.catalog_item_id:
                    catalog_item = session.get(FoodRegimenCatalogItem, item.catalog_item_id)
                    if catalog_item and catalog_item.item_type == "base_regimen":
                        base_ration_count += int(item.quantity)
            ration_count = base_ration_count or 1
            summaries[key]["special" if special else "standard"] += ration_count
            summaries[key]["total"] += ration_count
            rations.append(
                ProductionRationDetail(
                    admission_id=admission.id,
                    patient_name=patient_name,
                    service_name=service.name,
                    room_name=room.name,
                    bed_name=unit.label or unit.code,
                    meal_time=slot.meal_time,
                    ration_count=ration_count,
                    is_special=special,
                    items=item_labels,
                    instructions=slot.special_instructions or plan.general_instructions,
                    food_safety_alerts=alerts,
                )
            )
        for preparation in plan.modular_preparations:
            if meal_time and preparation.meal_time != meal_time:
                continue
            delivery = (
                MEAL_LABELS[preparation.meal_time]
                if preparation.meal_time
                else preparation.scheduled_time.strftime("%H:%M")
            )
            modular.append(
                ProductionModularDetail(
                    admission_id=admission.id,
                    patient_name=patient_name,
                    service_name=service.name,
                    room_name=room.name,
                    bed_name=unit.label or unit.code,
                    delivery=delivery,
                    product_name=preparation.product_name,
                    powder_grams=preparation.powder_grams,
                    diluent=preparation.diluent,
                    dilution_volume_ml=preparation.dilution_volume_ml,
                    units_per_delivery=preparation.units_per_delivery,
                    instructions=preparation.instructions,
                )
            )

    summary_rows = [
        ProductionSummaryLine(
            service_id=key[0],
            service_name=key[1],
            meal_time=key[2],
            standard_rations=value["standard"],
            special_rations=value["special"],
            total_rations=value["total"],
        )
        for key, value in summaries.items()
    ]
    preparation_rows = [
        ProductionPreparationLine(
            service_name=key[0],
            meal_time=key[1],
            item_name=key[2],
            unit=key[3],
            quantity=quantity,
            patient_count=len(item_patients[key]),
        )
        for key, quantity in item_totals.items()
    ]
    meal_order = {value: index for index, value in enumerate(MEAL_TIMES)}
    return ProductionConsolidatedRead(
        service_date=service_date,
        generated_at=utc_now(),
        meal_time=meal_time,
        summaries=sorted(summary_rows, key=lambda row: (row.service_name, meal_order[row.meal_time])),
        preparations=sorted(preparation_rows, key=lambda row: (row.service_name, meal_order[row.meal_time], row.item_name)),
        rations=sorted(rations, key=lambda row: (row.service_name, meal_order[row.meal_time], row.room_name, row.bed_name)),
        modular_preparations=sorted(modular, key=lambda row: (row.service_name, row.delivery, row.room_name, row.bed_name)),
        exceptions=sorted(exceptions, key=lambda row: (row.service_name or "", row.patient_name)),
    )


def _sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column]
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(value) for value in values) + 2, 55)


def production_xlsx(data: ProductionConsolidatedRead) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(
        workbook,
        "Resumen",
        ["Servicio", "Tiempo", "Estándar", "Especiales", "Total"],
        [[row.service_name, MEAL_LABELS[row.meal_time], row.standard_rations, row.special_rations, row.total_rations] for row in data.summaries],
    )
    _sheet(
        workbook,
        "Preparaciones",
        ["Servicio", "Tiempo", "Preparación", "Cantidad", "Unidad", "Pacientes"],
        [[row.service_name, MEAL_LABELS[row.meal_time], row.item_name, float(row.quantity), row.unit, row.patient_count] for row in data.preparations],
    )
    _sheet(
        workbook,
        "Raciones especiales",
        ["Paciente", "Servicio", "Sala", "Cama", "Tiempo", "Raciones", "Componentes", "Indicaciones", "Alertas alimentarias"],
        [[row.patient_name, row.service_name, row.room_name, row.bed_name, MEAL_LABELS[row.meal_time], row.ration_count, " | ".join(row.items), row.instructions or "", " | ".join(row.food_safety_alerts)] for row in data.rations if row.is_special],
    )
    _sheet(
        workbook,
        "Preparaciones_NE",
        ["Paciente", "Servicio", "Sala", "Cama", "Entrega", "Producto", "Polvo (g)", "Diluyente", "Volumen (mL)", "Vasos", "Indicaciones"],
        [[row.patient_name, row.service_name, row.room_name, row.bed_name, row.delivery, row.product_name, float(row.powder_grams), row.diluent, float(row.dilution_volume_ml), row.units_per_delivery, row.instructions or ""] for row in data.modular_preparations],
    )
    _sheet(
        workbook,
        "Control",
        ["Tipo", "Paciente", "Servicio", "Sala", "Cama", "Detalle"],
        [["Excepción", row.patient_name, row.service_name or "", row.room_name or "", row.bed_name or "", row.reason] for row in data.exceptions]
        + [["Generación", "", "", "", "", f"Fecha de servicio: {data.service_date.isoformat()} · Generado: {data.generated_at.isoformat()}"]],
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
