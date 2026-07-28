import hashlib
import re
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func
from sqlmodel import Session, select

from app.models.care_unit import CareUnit
from app.models.care_unit_layout_position import CareUnitLayoutPosition
from app.models.common import utc_now
from app.models.hospital_service import HospitalService
from app.models.room import Room
from app.services.audit_service import record_audit

EXPECTED_HEADERS = (
    "Servicio",
    "Sala o sector",
    "Tipo",
    "Camas/puestos",
    "Cantidad",
    "Observación",
)
UNIT_TYPE_MAP = {
    "cama": "bed",
    "camilla": "stretcher",
    "puesto": "station",
    "box": "box",
}
UNIT_TYPE_LABELS = {
    "bed": "Cama",
    "stretcher": "Camilla",
    "station": "Puesto",
    "box": "Box",
}
EMPTY_NOTES = {"", "-", "–", "—"}


class HospitalStructureImportError(ValueError):
    pass


@dataclass(frozen=True)
class HospitalImportRow:
    source_row: int
    service_name: str
    room_name: str
    unit_type: str
    unit_codes: tuple[str, ...]
    quantity: int
    notes: str | None


@dataclass
class ImportReport:
    source_name: str
    source_sha256: str
    rows: int
    services: int
    rooms: int
    care_units: int
    services_created: int = 0
    services_updated: int = 0
    rooms_created: int = 0
    rooms_updated: int = 0
    care_units_created: int = 0
    care_units_updated: int = 0
    layouts_created: int = 0
    layouts_updated: int = 0

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


def workbook_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _required_text(value: Any, *, row: int, column: str) -> str:
    normalized = " ".join(str(value).split()) if value is not None else ""
    if not normalized:
        raise HospitalStructureImportError(
            f"Fila {row}: la columna '{column}' no puede estar vacía."
        )
    return normalized


def _normalized_key(value: str) -> str:
    return " ".join(value.casefold().split())


def _normalize_type(value: Any, *, row: int) -> str:
    source_type = _required_text(value, row=row, column="Tipo")
    normalized = "".join(
        character
        for character in unicodedata.normalize("NFKD", source_type.casefold())
        if not unicodedata.combining(character)
    )
    try:
        return UNIT_TYPE_MAP[normalized]
    except KeyError as error:
        supported = ", ".join(UNIT_TYPE_MAP)
        raise HospitalStructureImportError(
            f"Fila {row}: tipo '{source_type}' no reconocido. Tipos admitidos: {supported}."
        ) from error


def _normalize_quantity(value: Any, *, row: int) -> int:
    if isinstance(value, bool):
        raise HospitalStructureImportError(f"Fila {row}: la cantidad debe ser un entero.")
    try:
        quantity = int(value)
    except (TypeError, ValueError) as error:
        raise HospitalStructureImportError(
            f"Fila {row}: la cantidad debe ser un entero."
        ) from error
    if isinstance(value, float) and not value.is_integer():
        raise HospitalStructureImportError(f"Fila {row}: la cantidad debe ser un entero.")
    if quantity < 1:
        raise HospitalStructureImportError(f"Fila {row}: la cantidad debe ser mayor que cero.")
    return quantity


def _normalize_unit_codes(value: Any, *, row: int) -> tuple[str, ...]:
    raw_codes = _required_text(value, row=row, column="Camas/puestos")
    codes = tuple(" ".join(part.split()).upper() for part in raw_codes.split(","))
    if any(not code for code in codes):
        raise HospitalStructureImportError(
            f"Fila {row}: la lista de camas/puestos contiene un código vacío."
        )
    if any(len(code) > 30 for code in codes):
        raise HospitalStructureImportError(
            f"Fila {row}: los códigos de camas/puestos no pueden superar 30 caracteres."
        )
    if len({_normalized_key(code) for code in codes}) != len(codes):
        raise HospitalStructureImportError(
            f"Fila {row}: la lista de camas/puestos contiene códigos duplicados."
        )
    return codes


def _normalize_notes(value: Any) -> str | None:
    if value is None:
        return None
    notes = " ".join(str(value).split())
    return None if notes in EMPTY_NOTES else notes


def read_hospital_workbook(path: Path) -> list[HospitalImportRow]:
    if not path.is_file():
        raise HospitalStructureImportError(f"No se encontró el archivo: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = tuple(
            " ".join(str(cell.value).split()) if cell.value is not None else ""
            for cell in sheet[1][:6]
        )
        if headers != EXPECTED_HEADERS:
            raise HospitalStructureImportError(
                "Las columnas del Excel no coinciden con el formato esperado: "
                + ", ".join(EXPECTED_HEADERS)
                + "."
            )

        rows: list[HospitalImportRow] = []
        seen_rooms: set[tuple[str, str]] = set()
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=2, max_col=6, values_only=True),
            start=2,
        ):
            if all(value is None or str(value).strip() == "" for value in values):
                continue
            service_name = _required_text(values[0], row=source_row, column="Servicio")
            room_name = _required_text(values[1], row=source_row, column="Sala o sector")
            if len(service_name) > 120 or len(room_name) > 120:
                raise HospitalStructureImportError(
                    f"Fila {source_row}: servicio y sala no pueden superar 120 caracteres."
                )
            room_key = (_normalized_key(service_name), _normalized_key(room_name))
            if room_key in seen_rooms:
                raise HospitalStructureImportError(
                    f"Fila {source_row}: la sala '{room_name}' está repetida en "
                    f"el servicio '{service_name}'."
                )
            seen_rooms.add(room_key)

            unit_codes = _normalize_unit_codes(values[3], row=source_row)
            quantity = _normalize_quantity(values[4], row=source_row)
            if len(unit_codes) != quantity:
                raise HospitalStructureImportError(
                    f"Fila {source_row}: Cantidad indica {quantity}, pero se encontraron "
                    f"{len(unit_codes)} códigos en Camas/puestos."
                )
            notes = _normalize_notes(values[5])
            if notes is not None and len(notes) > 500:
                raise HospitalStructureImportError(
                    f"Fila {source_row}: la observación no puede superar 500 caracteres."
                )
            rows.append(
                HospitalImportRow(
                    source_row=source_row,
                    service_name=service_name,
                    room_name=room_name,
                    unit_type=_normalize_type(values[2], row=source_row),
                    unit_codes=unit_codes,
                    quantity=quantity,
                    notes=notes,
                )
            )
    finally:
        workbook.close()

    if not rows:
        raise HospitalStructureImportError("El Excel no contiene filas para importar.")
    return rows


def _derive_code(value: str) -> str:
    ascii_value = "".join(
        character
        for character in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(character)
    )
    code = re.sub(r"[^A-Za-z0-9]+", "-", ascii_value).strip("-").upper()
    if not code:
        raise HospitalStructureImportError(f"No fue posible generar un código para '{value}'.")
    if len(code) <= 30:
        return code
    suffix = hashlib.sha1(value.encode("utf-8")).hexdigest()[:6].upper()
    return f"{code[:23].rstrip('-')}-{suffix}"


def _ensure_generated_codes_are_unique(rows: list[HospitalImportRow]) -> None:
    services: dict[str, str] = {}
    rooms: dict[tuple[str, str], str] = {}
    for row in rows:
        service_code = _derive_code(row.service_name)
        previous_service = services.setdefault(service_code, row.service_name)
        if _normalized_key(previous_service) != _normalized_key(row.service_name):
            raise HospitalStructureImportError(
                f"Los servicios '{previous_service}' y '{row.service_name}' generan "
                f"el mismo código '{service_code}'."
            )
        room_code = _derive_code(row.room_name)
        room_key = (service_code, room_code)
        previous_room = rooms.setdefault(room_key, row.room_name)
        if _normalized_key(previous_room) != _normalized_key(row.room_name):
            raise HospitalStructureImportError(
                f"Las salas '{previous_room}' y '{row.room_name}' generan "
                f"el mismo código '{room_code}' en '{row.service_name}'."
            )


def _set_values(instance: Any, values: dict[str, Any]) -> bool:
    changed = False
    for field, value in values.items():
        if getattr(instance, field) != value:
            setattr(instance, field, value)
            changed = True
    if changed:
        instance.updated_at = utc_now()
    return changed


def import_hospital_structure(
    session: Session,
    rows: list[HospitalImportRow],
    *,
    source_name: str,
    source_sha256: str,
) -> ImportReport:
    _ensure_generated_codes_are_unique(rows)
    service_names = list(dict.fromkeys(row.service_name for row in rows))
    report = ImportReport(
        source_name=source_name,
        source_sha256=source_sha256,
        rows=len(rows),
        services=len(service_names),
        rooms=len(rows),
        care_units=sum(row.quantity for row in rows),
    )

    services: dict[str, HospitalService] = {}
    for service_name in service_names:
        service_code = _derive_code(service_name)
        service = session.exec(
            select(HospitalService).where(
                func.lower(HospitalService.name) == service_name.lower()
            )
        ).first()
        code_owner = session.exec(
            select(HospitalService).where(
                func.lower(HospitalService.code) == service_code.lower()
            )
        ).first()
        if code_owner is not None and service is not None and code_owner.id != service.id:
            raise HospitalStructureImportError(
                f"El código generado '{service_code}' ya pertenece a otro servicio."
            )
        if code_owner is not None and service is None:
            if _normalized_key(code_owner.name) != _normalized_key(service_name):
                raise HospitalStructureImportError(
                    f"El código generado '{service_code}' ya pertenece a "
                    f"'{code_owner.name}'."
                )
            service = code_owner

        if service is None:
            service = HospitalService(code=service_code, name=service_name)
            session.add(service)
            session.flush()
            report.services_created += 1
        elif _set_values(service, {"name": service_name, "is_active": True}):
            session.add(service)
            report.services_updated += 1
        services[service_name] = service

    for row in rows:
        service = services[row.service_name]
        room_code = _derive_code(row.room_name)
        room = session.exec(
            select(Room).where(
                Room.service_id == service.id,
                func.lower(Room.code) == room_code.lower(),
            )
        ).first()
        if room is None:
            room = Room(
                service_id=service.id,
                code=room_code,
                name=row.room_name,
                notes=row.notes,
            )
            session.add(room)
            session.flush()
            report.rooms_created += 1
        elif _set_values(
            room,
            {
                "name": row.room_name,
                "notes": row.notes,
                "is_active": True,
            },
        ):
            session.add(room)
            report.rooms_updated += 1

        for position, unit_code in enumerate(row.unit_codes):
            care_unit = session.exec(
                select(CareUnit).where(
                    CareUnit.room_id == room.id,
                    func.lower(CareUnit.code) == unit_code.lower(),
                )
            ).first()
            care_unit_values = {
                "label": f"{UNIT_TYPE_LABELS[row.unit_type]} {unit_code}",
                "unit_type": row.unit_type,
                "is_active": True,
            }
            if care_unit is None:
                care_unit = CareUnit(
                    room_id=room.id,
                    code=unit_code,
                    **care_unit_values,
                )
                session.add(care_unit)
                session.flush()
                report.care_units_created += 1
            elif _set_values(care_unit, care_unit_values):
                session.add(care_unit)
                report.care_units_updated += 1

            layout = session.exec(
                select(CareUnitLayoutPosition).where(
                    CareUnitLayoutPosition.care_unit_id == care_unit.id
                )
            ).first()
            layout_values = {
                "grid_x": (position % 6) * 2,
                "grid_y": (position // 6) * 2,
                "width": 1,
                "height": 1,
            }
            if layout is None:
                session.add(
                    CareUnitLayoutPosition(
                        care_unit_id=care_unit.id,
                        **layout_values,
                    )
                )
                report.layouts_created += 1
            elif _set_values(layout, layout_values):
                session.add(layout)
                report.layouts_updated += 1

    record_audit(
        session,
        action="import",
        actor_user_id=None,
        entity_type="hospital_structure",
        after_state=report.as_dict(),
    )
    session.flush()
    return report
