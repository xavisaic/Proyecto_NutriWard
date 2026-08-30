from io import BytesIO

from openpyxl import load_workbook
from sqlmodel import select

from app.core.config import settings
from app.models.admission import Admission
from app.models.patient_location_history import PatientLocationHistory


MEAL_TIMES = (
    "breakfast",
    "morning_snack",
    "lunch",
    "afternoon_snack",
    "dinner",
    "night_snack",
)


def authenticate(client, role: str = "nutricionista") -> dict:
    response = client.post(
        "/api/v1/auth/login",
        json={
            "email": f"{role}@nutriward.local",
            "password": settings.demo_user_password,
        },
    )
    assert response.status_code == 200
    return response.json()


def located_admission(db_session) -> Admission:
    admission = db_session.exec(
        select(Admission)
        .join(PatientLocationHistory, PatientLocationHistory.admission_id == Admission.id)
        .where(Admission.status == "active", PatientLocationHistory.ended_at.is_(None))
        .order_by(Admission.admitted_at)
    ).first()
    assert admission is not None
    return admission


def plan_payload(catalog_item_id: str) -> dict:
    slots = [
        {
            "meal_time": meal,
            "fulfillment_status": "not_applicable",
            "items": [],
        }
        for meal in MEAL_TIMES
    ]
    slots[0] = {
        "meal_time": "breakfast",
        "fulfillment_status": "ordered",
        "is_special": True,
        "special_instructions": "Enviar solamente medio pan.",
        "items": [
            {
                "catalog_item_id": catalog_item_id,
                "quantity": 1,
                "unit": "porción",
            },
            {
                "custom_name": "Pan sin agregado",
                "quantity": 0.5,
                "unit": "unidad",
            },
            {
                "custom_name": "Galletas de soda",
                "quantity": 7,
                "unit": "unidad",
            },
        ],
    }
    return {
        "effective_from": "2026-08-29",
        "validity_mode": "until_changed",
        "oral_enabled": True,
        "enteral_enabled": True,
        "parenteral_enabled": False,
        "general_instructions": "Sin gluten.",
        "slots": slots,
        "modular_preparations": [
            {
                "preparation_type": "protein_bolus",
                "product_name": "Módulo proteico",
                "powder_grams": 10,
                "diluent": "Agua",
                "dilution_volume_ml": 80,
                "units_per_delivery": 1,
                "meal_time": "morning_snack",
            }
        ],
    }


def test_catalog_permissions_seed_and_openapi(client) -> None:
    assert client.get("/api/v1/food-regimen-catalog").status_code == 401
    authenticate(client, "administrador")
    assert client.get("/api/v1/food-regimen-catalog").status_code == 403

    client.cookies.clear()
    authenticate(client, "alimentacion")
    active = client.get("/api/v1/food-regimen-catalog")
    assert active.status_code == 200
    assert any(row["display_name"] == "Leche con 3 medida de espesante" for row in active.json())
    all_items = client.get("/api/v1/food-regimen-catalog?include_inactive=true")
    assert len(all_items.json()) == 143
    assert any(row["display_name"] == "Ayunas" and not row["is_active"] for row in all_items.json())
    indicated = next(row for row in all_items.json() if row["display_name"] == "INDICADO s/s+1")
    assert indicated["standard_recipe_note"] == (
        "Sin sal. Entregar 1 g de sal aparte. Régimen diabético."
    )
    unsupplemented = next(
        row for row in all_items.json() if row["display_name"] == "Papilla NO suple"
    )
    assert unsupplemented["standard_recipe_note"] == "Papilla sin suplementación estándar."

    schema = client.get("/openapi.json").json()
    assert "/api/v1/admissions/{admission_id}/meal-plans" in schema["paths"]
    assert "/api/v1/food-production/consolidated.xlsx" in schema["paths"]


def test_combined_oral_enteral_plan_production_and_xlsx(client, db_session) -> None:
    admission = located_admission(db_session)
    auth = authenticate(client)
    headers = {"X-CSRF-Token": auth["csrf_token"]}
    catalog = client.get("/api/v1/food-regimen-catalog?search=liviano").json()
    liviano = next(row for row in catalog if row["display_name"] == "Liviano")
    invalid_payload = plan_payload(liviano["id"])
    invalid_payload["slots"][0]["items"][0]["quantity"] = 1.5
    invalid = client.post(
        f"/api/v1/admissions/{admission.id}/meal-plans",
        json=invalid_payload,
        headers=headers,
    )
    assert invalid.status_code == 422

    payload = plan_payload(liviano["id"])
    payload["slots"][0]["items"][0]["quantity"] = 2

    created = client.post(
        f"/api/v1/admissions/{admission.id}/meal-plans",
        json=payload,
        headers=headers,
    )
    assert created.status_code == 201, created.text
    body = created.json()
    assert body["status"] == "draft"
    assert body["oral_enabled"] and body["enteral_enabled"]
    assert [slot["meal_time"] for slot in body["slots"]] == list(MEAL_TIMES)
    assert body["modular_preparations"][0]["powder_grams"] == "10.000"

    stale = client.put(
        f"/api/v1/meal-plans/{body['id']}",
        json={**payload, "version": 99},
        headers=headers,
    )
    assert stale.status_code == 409

    finalized = client.post(
        f"/api/v1/meal-plans/{body['id']}/finalize",
        json={"version": 1},
        headers=headers,
    )
    assert finalized.status_code == 200, finalized.text
    assert finalized.json()["status"] == "finalized"

    client.cookies.clear()
    authenticate(client, "alimentacion")
    consolidated = client.get(
        "/api/v1/food-production/consolidated",
        params={"service_date": "2026-08-29"},
    )
    assert consolidated.status_code == 200, consolidated.text
    data = consolidated.json()
    ration = next(row for row in data["rations"] if row["admission_id"] == str(admission.id))
    assert ration["meal_time"] == "breakfast"
    assert ration["ration_count"] == 2
    assert ration["is_special"] is True
    assert any("Galletas de soda" in item for item in ration["items"])
    bolus = next(
        row for row in data["modular_preparations"] if row["admission_id"] == str(admission.id)
    )
    assert bolus["powder_grams"] == "10.000"
    assert bolus["dilution_volume_ml"] == "80.000"
    summary = next(
        row
        for row in data["summaries"]
        if row["service_name"] == ration["service_name"] and row["meal_time"] == "breakfast"
    )
    assert summary["special_rations"] == 2
    assert summary["total_rations"] == 2
    liviano_total = next(
        row
        for row in data["preparations"]
        if row["service_name"] == ration["service_name"]
        and row["meal_time"] == "breakfast"
        and row["item_name"] == "Liviano"
    )
    assert liviano_total["quantity"] == "2.000"

    exported = client.get(
        "/api/v1/food-production/consolidated.xlsx",
        params={"service_date": "2026-08-29"},
    )
    assert exported.status_code == 200
    assert exported.content[:2] == b"PK"
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    assert workbook.sheetnames == [
        "Resumen",
        "Preparaciones",
        "Raciones especiales",
        "Preparaciones_NE",
        "Control",
    ]


def test_enteral_only_bolus_does_not_create_tray(client, db_session) -> None:
    admission = located_admission(db_session)
    auth = authenticate(client)
    headers = {"X-CSRF-Token": auth["csrf_token"]}
    payload = plan_payload(client.get("/api/v1/food-regimen-catalog?search=jalea").json()[0]["id"])
    payload["oral_enabled"] = False
    payload["slots"] = [
        {"meal_time": meal, "fulfillment_status": "not_applicable", "items": []}
        for meal in MEAL_TIMES
    ]
    created = client.post(
        f"/api/v1/admissions/{admission.id}/meal-plans",
        json=payload,
        headers=headers,
    )
    assert created.status_code == 201, created.text
    body = created.json()
    assert client.post(
        f"/api/v1/meal-plans/{body['id']}/finalize",
        json={"version": body["version"]},
        headers=headers,
    ).status_code == 200

    client.cookies.clear()
    authenticate(client, "alimentacion")
    data = client.get(
        "/api/v1/food-production/consolidated",
        params={"service_date": "2026-08-29"},
    ).json()
    assert not any(row["admission_id"] == str(admission.id) for row in data["rations"])
    assert any(row["admission_id"] == str(admission.id) for row in data["modular_preparations"])
